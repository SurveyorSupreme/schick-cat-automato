import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from copy import copy
from io import BytesIO
import re

# --- 1. THE CORE ENGINE ---
def process_cat_sheet(template_file, csv_files):
    # Load with keep_vba=True is critical for CCC macro-enabled sheets
    wb = openpyxl.load_workbook(template_file, keep_vba=True)
    
    ft_sheet = wb['Feature_Templates']
    routing_map = {}
    master_row_map = {} # Stores the row index for each code
    current_cat = None
    
    # 1. Build a map of all codes and their "Golden Row" index
    for r in range(1, ft_sheet.max_row + 1):
        val = str(ft_sheet.cell(row=r, column=1).value or "").strip()
        if "Point" in val: current_cat = "Point Asset Inputs"
        elif "Line" in val: current_cat = "Line Asset Inputs"
        elif "Polygon" in val: current_cat = "Polygon Asset Inputs"
        
        if re.match(r'^[A-Z]\d{2}$', val):
            routing_map[val] = current_cat
            master_row_map[val] = r

    # 2. Process the CSVs
    for csv_file in csv_files:
        df = pd.read_csv(csv_file, header=None)
        
        for _, csv_row in df.iterrows():
            code = str(csv_row[0]).strip()
            target_sheet_name = routing_map.get(code)
            
            if target_sheet_name and target_sheet_name in wb.sheetnames:
                target_sheet = wb[target_sheet_name]
                source_row_idx = master_row_map[code]
                
                # Find the next empty row in the target sheet
                next_row = 2
                while target_sheet.cell(row=next_row, column=1).value is not None:
                    next_row += 1
                
                # --- STEP A: CLONE THE ROW (Styles & Values) ---
                for col in range(1, ft_sheet.max_column + 1):
                    source_cell = ft_sheet.cell(row=source_row_idx, column=col)
                    target_cell = target_sheet.cell(row=next_row, column=col)
                    
                    target_cell.value = source_cell.value
                    if source_cell.has_style:
                        target_cell.font = copy(source_cell.font)
                        target_cell.border = copy(source_cell.border)
                        target_cell.fill = copy(source_cell.fill)
                        target_cell.number_format = copy(source_cell.number_format)
                        target_cell.alignment = copy(source_cell.alignment)
                
                # --- STEP B: MIGRATE DROPDOWNS (Data Validation) ---
                # We look at every dropdown rule in the template sheet
                for dv in ft_sheet.data_validations.dataValidation:
                    # Check if this rule applied to our master row
                    # We check if the cell (e.g., 'B211') is in the validation range
                    affected_cells = dv.cells
                    # Simplify: if the rule is column-wide or covers our row
                    source_cell_coord = f"{get_column_letter(1)}{source_row_idx}" 
                    # (Note: This is a simplification; we check for column intersection)
                    
                    # If the DV formula exists, we re-apply it to the target cell in the same column
                    # We iterate columns to find which ones have dropdowns
                    for col in range(1, ft_sheet.max_column + 1):
                        cell_ref = f"{get_column_letter(col)}{source_row_idx}"
                        if cell_ref in dv:
                            # Create a matching validation in the target sheet
                            new_dv = DataValidation(
                                type=dv.type, formula1=dv.formula1, formula2=dv.formula2,
                                allow_blank=dv.allow_blank, showErrorMessage=dv.showErrorMessage,
                                showInputMessage=dv.showInputMessage
                            )
                            target_sheet.add_data_validation(new_dv)
                            new_dv.add(target_sheet.cell(row=next_row, column=col))

                # --- STEP C: OVERWRITE WITH CSV VALUES ---
                # This ensures we keep the dropdowns we just pasted but put in the real data
                for col_idx, value in enumerate(csv_row):
                    if pd.notna(value):
                        target_sheet.cell(row=next_row, column=col_idx + 1).value = value

    # Save to buffer
    out_buffer = BytesIO()
    wb.save(out_buffer)
    out_buffer.seek(0)
    return out_buffer

# --- 2. STREAMLIT UI ---
st.set_page_config(page_title="Schick CAT | Deep Clone", page_icon="🚜")
st.title("🚜 Schick Group: Master Row Cloner")
st.markdown("Clones **Feature_Templates** rows (including dropdowns) into the Input sheets.")

template = st.file_uploader("1. Upload CCC Template (.xlsm)", type=['xlsm'])
csvs = st.file_uploader("2. Upload 12d CSVs", type=['csv'], accept_multiple_files=True)

if st.button("Generate Validated CAT"):
    if template and csvs:
        with st.spinner("Deep cloning council rules..."):
            try:
                output = process_cat_sheet(template, csvs)
                st.success("Done! Your survey data is now inside official Council rows.")
                st.download_button("📥 Download for CCC Portal", output, file_name="Schick_Validated_CAT.xlsm")
            except Exception as e:
                st.error(f"Error: {e}")
